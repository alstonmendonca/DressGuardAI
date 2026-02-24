from fastapi import FastAPI, UploadFile, File, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
import numpy as np
import cv2
from detector import DressDetector
from utils.compliance import is_compliant, ComplianceManager
from utils.logger import setup_logging
from utils.cache import get_cache
from utils.model_discovery import get_model_discovery
from utils.visibility_checker import get_visibility_checker
from utils.violation_logger import get_violation_logger
from utils.face_recognition_insightface import detect_and_identify_faces
from utils.whatsapp_sender import get_whatsapp_sender
from config import (MODELS_FOLDER, WEBCAM_DETECTION_INTERVAL,
                    WEBCAM_JPEG_QUALITY, WEBCAM_FPS_LIMIT, 
                    WEBCAM_SKIP_FRAMES)
import logging
from typing import Optional, List
from datetime import datetime
import os
from contextlib import asynccontextmanager
import asyncio
import time
import subprocess
import sys

# Setup logging with rotating file handler
log_level = os.getenv("LOG_LEVEL", "INFO")
log_file = os.getenv("LOG_FILE", "logs/dressguard.log")
setup_logging(log_level=log_level, log_file=log_file)

logger = logging.getLogger(__name__)

# Global variable to track WhatsApp service process
whatsapp_process = None

# Background task for cache cleanup
async def cleanup_task():
    """Periodic cache cleanup task"""
    cache = get_cache()
    while True:
        await asyncio.sleep(300)  # Run every 5 minutes
        try:
            cache.cleanup_expired()
        except Exception as e:
            logger.error(f"Cache cleanup error: {e}")

def start_whatsapp_service():
    """Start the WhatsApp Web service in the background"""
    global whatsapp_process
    try:
        whatsapp_service_dir = os.path.join(os.path.dirname(__file__), "whatsapp-service")
        
        if not os.path.exists(whatsapp_service_dir):
            logger.warning(f"WhatsApp service directory not found: {whatsapp_service_dir}")
            return None
            
        server_js = os.path.join(whatsapp_service_dir, "server.js")
        if not os.path.exists(server_js):
            logger.warning(f"WhatsApp service server.js not found: {server_js}")
            return None
        
        logger.info("Starting WhatsApp Web service...")
        
        # Start Node.js service as background process
        if sys.platform == "win32":
            # Windows: Start with CREATE_NEW_CONSOLE to run independently
            whatsapp_process = subprocess.Popen(
                ["node", "server.js"],
                cwd=whatsapp_service_dir,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                creationflags=subprocess.CREATE_NEW_CONSOLE
            )
        else:
            # Linux/Mac: Start with nohup-like behavior
            whatsapp_process = subprocess.Popen(
                ["node", "server.js"],
                cwd=whatsapp_service_dir,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                start_new_session=True
            )
        
        logger.info(f"WhatsApp service started with PID: {whatsapp_process.pid}")
        logger.info("WhatsApp service will show QR code in its own console window")
        logger.info("Scan the QR code with your phone to connect WhatsApp")
        return whatsapp_process
        
    except FileNotFoundError:
        logger.error("Node.js not found. Please install Node.js to use WhatsApp service")
        return None
    except Exception as e:
        logger.error(f"Failed to start WhatsApp service: {e}")
        return None

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifespan context manager for startup and shutdown events"""
    global whatsapp_process
    
    # Startup
    logger.info("Starting DressGuard API...")
    cleanup_task_handle = asyncio.create_task(cleanup_task())
    
    # Start WhatsApp service
    whatsapp_process = start_whatsapp_service()
    
    yield
    
    # Shutdown
    logger.info("Shutting down DressGuard API...")
    cleanup_task_handle.cancel()
    try:
        await cleanup_task_handle
    except asyncio.CancelledError:
        pass
    
    # Stop WhatsApp service
    if whatsapp_process:
        logger.info("Stopping WhatsApp service...")
        try:
            whatsapp_process.terminate()
            whatsapp_process.wait(timeout=5)
            logger.info("WhatsApp service stopped")
        except Exception as e:
            logger.error(f"Error stopping WhatsApp service: {e}")
            try:
                whatsapp_process.kill()
            except:
                pass

app = FastAPI(
    title="DressGuard API",
    description="AI-powered clothing compliance detection system",
    version="1.0.0",
    lifespan=lifespan
)

# Initialize detector with error handling
try:
    detector = DressDetector()
    logger.info("DressDetector initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize DressDetector: {e}")
    detector = None

# Initialize compliance manager
compliance_manager = ComplianceManager()

# Initialize violation logger
violation_logger = get_violation_logger()

# Initialize WhatsApp sender
whatsapp_sender = get_whatsapp_sender()

# Global face detection toggle (OFF by default)
face_detection_enabled = False

# Store last uploaded image detection for manual logging
last_upload_detection = None

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # for dev, restrict in prod
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration constants
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "name": "DressGuard API",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "detect": "/detect/",
            "models": "/models/",
            "health": "/health/",
            "switch_model": "/switch-model/",
            "current_model": "/current-model/"
        }
    }

@app.get("/health/")
async def health_check():
    """Health check endpoint for monitoring"""
    if detector is None:
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "detector": "not initialized",
                "message": "Detector failed to initialize"
            }
        )
    
    model_discovery = get_model_discovery(MODELS_FOLDER)
    available_models = model_discovery.get_all_models()
    
    return {
        "status": "healthy",
        "detector": "initialized",
        "current_model": detector.current_model,
        "available_models": list(available_models.keys())
    }

@app.get("/device/")
async def get_device_info():
    """Get information about the device being used for inference (GPU/CPU)"""
    if detector is None:
        return JSONResponse(
            status_code=503,
            content={"error": "Detector not initialized"}
        )
    
    device_info = detector.get_device_info()
    
    # Add face detection device info
    try:
        from utils.face_recognition_insightface import get_face_app
        face_app = get_face_app()
        if face_app:
            # Check if using GPU or CPU
            face_device = "CPU"
            try:
                # InsightFace uses ONNX Runtime providers
                import onnxruntime as ort
                providers = ort.get_available_providers()
                if 'CUDAExecutionProvider' in providers:
                    face_device = "GPU (CUDA)"
                elif 'CPUExecutionProvider' in providers:
                    face_device = "CPU"
            except:
                face_device = "CPU"
            
            device_info["face_detection_device"] = face_device
        else:
            device_info["face_detection_device"] = "Not initialized"
    except ImportError:
        # Fallback to old face_recognition (always CPU)
        device_info["face_detection_device"] = "CPU (dlib)"
    
    return device_info

@app.get("/models/")
async def get_available_models():
    """Get list of available models with their metadata (dynamically discovered)"""
    try:
        model_discovery = get_model_discovery(MODELS_FOLDER)
        models = model_discovery.get_all_models()
        
        models_info = []
        for model_id, model_data in models.items():
            models_info.append({
                "id": model_id,
                "path": model_data["path"],
                "classes": model_data["classes"],
                "class_count": model_data["class_count"],
                "is_current": model_id == detector.current_model if detector else False
            })
        
        return {
            "models": models_info,
            "current_model": detector.current_model if detector else None,
            "total": len(models_info)
        }
    except Exception as e:
        logger.error(f"Error fetching models: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch models: {str(e)}")

@app.post("/detect/")
async def detect_dress(file: UploadFile = File(...), model: Optional[str] = None):
    """Detect clothing items in uploaded image"""
    if detector is None:
        raise HTTPException(status_code=503, detail="Detector not initialized")
    
    try:
        # Validate file extension
        file_ext = file.filename.split('.')[-1].lower()
        if f".{file_ext}" not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid file type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
            )
        
        # Read and validate file size
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"File too large. Maximum size: {MAX_FILE_SIZE / (1024*1024)}MB"
            )
        
        # Decode image
        nparr = np.frombuffer(content, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if image is None:
            raise HTTPException(status_code=400, detail="Failed to decode image")
        
        # Validate image dimensions
        h, w = image.shape[:2]
        if w < 50 or h < 50:
            raise HTTPException(status_code=400, detail="Image too small (minimum 50x50 pixels)")
        if w > 4096 or h > 4096:
            raise HTTPException(status_code=400, detail="Image too large (maximum 4096x4096 pixels)")

        # Switch model if specified
        if model and model != detector.current_model:
            success = detector.switch_model(model)
            if not success:
                logger.warning(f"Failed to switch to model: {model}")

        # Perform detection
        detected_clothes = detector.detect(image)
        
        # Check compliance using the compliance manager
        compliant, non_compliant_items, compliance_details = compliance_manager.check_compliance(detected_clothes)
        
        logger.info(f"Detection complete: {len(detected_clothes)} items found, compliant: {compliant}")

        # Initialize face detection results
        face_results = []
        violation_logged = False
        
        # Always perform face detection for uploads if enabled
        if face_detection_enabled:
            try:
                # Convert to RGB for face detection
                image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                face_results = detect_and_identify_faces(image_rgb)
                
                if face_results:
                    detected_names = [f"{face.get('name')} ({face.get('confidence', 0):.1f}%)" for face in face_results]
                    logger.info(f"Face detection results for uploaded image: {detected_names}")
                    
                    # Check for multiple people
                    if len(face_results) > 1:
                        logger.warning(f"Multiple people detected in uploaded image: {detected_names}")
                else:
                    logger.info("No faces detected in uploaded image - marking as Unknown")
                    # Create Unknown face entry for consistent logging
                    face_results = [{'name': 'Unknown', 'confidence': 0, 'user_id': None}]
            except Exception as e:
                logger.error(f"Error during face detection for uploaded image: {e}", exc_info=True)
                # On error, mark as Unknown
                face_results = [{'name': 'Unknown', 'confidence': 0, 'user_id': None}]
        else:
            # Face detection disabled - mark as Unknown
            logger.info("Face detection disabled - marking as Unknown if logged")
            face_results = [{'name': 'Unknown', 'confidence': 0, 'user_id': None}]
        
        # Store upload detection for manual logging (if non-compliant)
        global last_upload_detection
        if not compliant:
            last_upload_detection = {
                'image': image.copy(),
                'detected_clothes': detected_clothes,
                'face_results': face_results,
                'compliance_info': {
                    'is_compliant': compliant,
                    'non_compliant_items': non_compliant_items
                },
                'model': detector.current_model,
                'timestamp': time.time()
            }
            logger.info("Stored upload detection for manual logging via 'Log Image Result' button")
        else:
            logger.info("✓ Upload is compliant - no violation to log")
            last_upload_detection = None

        return {
            "clothes_detected": detected_clothes,
            "image_width": w,
            "image_height": h,
            "compliant": compliant,
            "non_compliant_items": non_compliant_items,
            "compliance_details": compliance_details,
            "model_used": detector.current_model,
            "total_detections": len(detected_clothes),
            "faces_detected": len(face_results),
            "violation_logged": violation_logged,
            "can_log_image": not compliant  # Show Log Image Result button if non-compliant
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Detection error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Detection failed: {str(e)}")

@app.post("/log-upload-image/")
async def log_upload_image():
    """
    Manually log the last uploaded non-compliant image.
    This is separate from video stream logging.
    """
    global last_upload_detection
    
    if last_upload_detection is None:
        raise HTTPException(
            status_code=404,
            detail="No upload available to log. Please upload a non-compliant image first."
        )
    
    try:
        # Check if upload is still recent (within 10 minutes)
        if time.time() - last_upload_detection['timestamp'] > 600:
            last_upload_detection = None
            raise HTTPException(
                status_code=410,
                detail="Upload expired (10 min timeout). Please re-upload the image."
            )
        
        logger.info("=== MANUAL IMAGE LOGGING START ===")
        logger.info(f"Logging enabled: {violation_logger.is_logging_enabled()}")
        logger.info(f"Non-compliant items: {last_upload_detection['compliance_info']['non_compliant_items']}")
        logger.info(f"Face results: {[(f['name'], f.get('confidence', 0)) for f in last_upload_detection['face_results']]}")
        logger.info(f"Model: {last_upload_detection['model']}")
        logger.info("Calling save_violation with skip_cooldown=True (bypasses logging enabled check)")
        
        # Log the violation with all data (including license plates for Vehicle Helmet model)
        # Skip cooldown for manual image uploads (bypasses both cooldown AND logging_enabled check)
        violation_logged = violation_logger.save_violation(
            last_upload_detection['image'].copy(),
            last_upload_detection['detected_clothes'],
            last_upload_detection['face_results'],
            last_upload_detection['compliance_info'],
            current_model=last_upload_detection['model'],
            skip_cooldown=True
        )
        
        if violation_logged:
            face_names = [f.get('name') for f in last_upload_detection['face_results']]
            logger.info(f"✓ Image violation SUCCESSFULLY logged: {face_names}")
            logger.info(f"✓ Check dashboard for today's date: {datetime.now().strftime('%Y-%m-%d')}")
            logger.info("=== MANUAL IMAGE LOGGING SUCCESS ===")
            
            # Clear stored detection after successful logging
            last_upload_detection = None
            
            return {
                "success": True,
                "message": "Image violation logged successfully",
                "faces": face_names,
                "date": datetime.now().strftime('%Y-%m-%d')
            }
        else:
            logger.error("✗ Image violation NOT logged - save_violation returned False")
            logger.error("This should NOT happen with skip_cooldown=True!")
            logger.error(f"Logging enabled: {violation_logger.is_logging_enabled()}")
            logger.error(f"Pending tasks: {violation_logger.pending_tasks}/{violation_logger.max_pending_tasks}")
            logger.error("=== MANUAL IMAGE LOGGING FAILED ===")
            raise HTTPException(
                status_code=500,
                detail="Failed to log image. Check server logs for details."
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error during manual image logging: {e}", exc_info=True)
        logger.error("=== MANUAL IMAGE LOGGING ERROR ===")
        raise HTTPException(status_code=500, detail=f"Image logging failed: {str(e)}")

@app.get("/log-upload-image/status/")
async def get_upload_log_status():
    """Check if there's an upload available to log"""
    global last_upload_detection
    
    if last_upload_detection is None:
        return {
            "available": False,
            "message": "No upload available"
        }
    
    # Check if expired
    age = time.time() - last_upload_detection['timestamp']
    if age > 600:
        last_upload_detection = None
        return {
            "available": False,
            "message": "Upload expired"
        }
    
    return {
        "available": True,
        "message": "Upload ready to log",
        "age_seconds": int(age),
        "expires_in_seconds": int(600 - age),
        "non_compliant_items": last_upload_detection['compliance_info']['non_compliant_items'],
        "faces": [f.get('name') for f in last_upload_detection['face_results']]
    }

@app.post("/switch-model/")
async def switch_model(model_name: str = Body(..., embed=True)):
    """Switch to a different detection model (dynamically discovered)"""
    if detector is None:
        raise HTTPException(status_code=503, detail="Detector not initialized")
    
    try:
        logger.info(f"Received model switch request: '{model_name}'")
        
        # Check if model exists using model discovery
        model_discovery = get_model_discovery(MODELS_FOLDER)
        available_models = model_discovery.get_all_models()
        
        # Find model with case-insensitive match
        matched_model_id = None
        for model_id in available_models.keys():
            if model_id.lower() == model_name.lower():
                matched_model_id = model_id
                break
        
        if not matched_model_id:
            logger.warning(f"Model '{model_name}' not found in available models")
            raise HTTPException(
                status_code=400,
                detail=f"Invalid model '{model_name}'. Available models: {', '.join(available_models.keys())}"
            )
        
        logger.info(f"Matched model: '{model_name}' -> '{matched_model_id}'")
        success = detector.switch_model(matched_model_id)
        
        if success:
            # Keep the same compliance configuration (don't reload model-specific config)
            # This ensures compliance settings remain stable across model switches
            logger.info(f"Successfully switched to model: {matched_model_id}")
            logger.info(f"Using current compliance config (stable across models)")
            return {
                "success": True,
                "current_model": detector.current_model,
                "message": f"Successfully switched to {detector.current_model}"
            }
        else:
            logger.error(f"Failed to switch to model: {matched_model_id}")
            raise HTTPException(
                status_code=500,
                detail="Model switch failed. Check server logs for details."
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error switching model: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Model switch error: {str(e)}")

@app.get("/current-model/")
async def get_current_model():
    """Get the currently active model"""
    if detector is None:
        raise HTTPException(status_code=503, detail="Detector not initialized")
    
    # Get model info from dynamic discovery
    model_discovery = get_model_discovery(MODELS_FOLDER)
    model_info = model_discovery.get_model_info(detector.current_model)
    
    return {
        "current_model": detector.current_model,
        "model_info": model_info if model_info else {}
    }

@app.get("/stats/")
async def get_stats():
    """Get system statistics including cache and performance metrics"""
    cache = get_cache()
    cache_stats = cache.get_stats()
    cache_size = cache.get_size_estimate()
    
    return {
        "cache": {
            **cache_stats,
            "size_bytes": cache_size,
            "size_mb": round(cache_size / (1024 * 1024), 2)
        },
        "detector": {
            "initialized": detector is not None,
            "current_model": detector.current_model if detector else None,
            "available_models": len(model_discovery.get_all_models())
        }
    }

@app.post("/cache/clear/")
async def clear_cache():
    """Clear the API cache"""
    try:
        cache = get_cache()
        cache.clear()
        logger.info("Cache cleared via API request")
        return {
            "success": True,
            "message": "Cache cleared successfully"
        }
    except Exception as e:
        logger.error(f"Error clearing cache: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# Compliance Configuration Endpoints
# ============================================================================

@app.get("/compliance/config/")
async def get_compliance_config():
    """Get current compliance configuration for the active model"""
    config = compliance_manager.get_config()
    
    # Add current model information
    config["current_model"] = detector.current_model if detector else None
    
    # Filter to only include classes from the current model
    if detector and hasattr(detector.model, 'names'):
        current_model_classes = set(c.lower() for c in detector.model.names.values())
        
        # Filter compliant and non-compliant classes to show only those available in current model
        config["compliant_classes"] = [
            c for c in config["compliant_classes"] 
            if c in current_model_classes
        ]
        config["non_compliant_classes"] = [
            c for c in config["non_compliant_classes"] 
            if c in current_model_classes
        ]
        
        # Add info about filtered classes (classes in config but not in current model)
        all_compliant = set(compliance_manager.compliant_classes)
        all_non_compliant = set(compliance_manager.non_compliant_classes)
        
        config["filtered_out"] = {
            "compliant": sorted(list(all_compliant - current_model_classes)),
            "non_compliant": sorted(list(all_non_compliant - current_model_classes))
        }
    
    return config

@app.post("/compliance/config/")
async def update_compliance_config(
    compliant_classes: List[str] = Body(...),
    non_compliant_classes: List[str] = Body(...),
    min_confidence: Optional[float] = Body(0.5)
):
    """
    Update compliance configuration
    
    Args:
        compliant_classes: List of approved clothing items
        non_compliant_classes: List of prohibited clothing items
        min_confidence: Minimum confidence threshold (0-1)
    """
    try:
        # Normalize all class names
        compliant_set = set(c.lower().strip() for c in compliant_classes)
        non_compliant_set = set(c.lower().strip() for c in non_compliant_classes)
        
        # Remove any duplicates - if a class is in both lists, keep it in non-compliant
        # This ensures safety: we'd rather flag something as non-compliant than miss it
        overlap = compliant_set & non_compliant_set
        if overlap:
            logger.warning(f"Classes in both lists (keeping in non-compliant): {overlap}")
            compliant_set -= overlap
        
        # Update the configuration
        compliance_manager.compliant_classes = compliant_set
        compliance_manager.non_compliant_classes = non_compliant_set
        compliance_manager.min_confidence = min_confidence
        
        # Save with current model name for model-specific config
        model_name = detector.current_model if detector else None
        compliance_manager.save_config(model_name)
        
        logger.info(f"Compliance config updated: {len(compliant_set)} compliant, "
                   f"{len(non_compliant_set)} non-compliant, "
                   f"{len(overlap)} duplicates removed")
        
        return {
            "success": True,
            "message": "Compliance configuration updated",
            "duplicates_removed": list(overlap) if overlap else [],
            "config": compliance_manager.get_config()
        }
    except Exception as e:
        logger.error(f"Error updating compliance config: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/compliance/add-compliant/")
async def add_compliant_class(class_name: str = Body(..., embed=True)):
    """Add a class to the compliant list"""
    try:
        model_name = detector.current_model if detector else None
        compliance_manager.add_compliant_class(class_name, model_name)
        return {
            "success": True,
            "message": f"Added '{class_name}' to compliant classes",
            "config": compliance_manager.get_config()
        }
    except Exception as e:
        logger.error(f"Error adding compliant class: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/compliance/add-non-compliant/")
async def add_non_compliant_class(class_name: str = Body(..., embed=True)):
    """Add a class to the non-compliant list"""
    try:
        model_name = detector.current_model if detector else None
        compliance_manager.add_non_compliant_class(class_name, model_name)
        return {
            "success": True,
            "message": f"Added '{class_name}' to non-compliant classes",
            "config": compliance_manager.get_config()
        }
    except Exception as e:
        logger.error(f"Error adding non-compliant class: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/compliance/remove-class/")
async def remove_class(class_name: str = Body(..., embed=True)):
    """Remove a class from both compliant and non-compliant lists"""
    try:
        model_name = detector.current_model if detector else None
        compliance_manager.remove_class(class_name, model_name)
        return {
            "success": True,
            "message": f"Removed '{class_name}' from compliance lists",
            "config": compliance_manager.get_config()
        }
    except Exception as e:
        logger.error(f"Error removing class: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/compliance/detected-classes/")
async def get_all_detected_classes():
    """Get list of all unique classes across all models or current model"""
    if detector is None:
        raise HTTPException(status_code=503, detail="Detector not initialized")
    
    try:
        # Get model discovery instance
        model_discovery = get_model_discovery(MODELS_FOLDER)
        
        # Get all unique classes across all models
        all_classes = model_discovery.get_all_unique_classes()
        
        # Get classes from current model
        current_model_classes = []
        if hasattr(detector.model, 'names'):
            current_model_classes = list(detector.model.names.values())
        
        return {
            "classes": sorted(all_classes),  # All unique classes across all models
            "current_model_classes": sorted(current_model_classes),  # Current model only
            "count": len(all_classes),
            "current_model": detector.current_model
        }
    except Exception as e:
        logger.error(f"Error getting detected classes: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# Violation Logging Endpoints
# ============================================================================

@app.post("/logging/toggle/")
async def toggle_logging():
    """Toggle violation logging on/off"""
    try:
        new_state = violation_logger.toggle_logging()
        return {
            "success": True,
            "logging_enabled": new_state,
            "message": f"Logging {'enabled' if new_state else 'disabled'}"
        }
    except Exception as e:
        logger.error(f"Error toggling logging: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/logging/status/")
async def get_logging_status():
    """Get current logging status with statistics"""
    stats = violation_logger.get_stats()
    return {
        "logging_enabled": violation_logger.is_logging_enabled(),
        "cooldown_seconds": stats["cooldown_seconds"],
        "active_violations": stats["active_violations"],
        "log_folder": stats["log_folder"]
    }

@app.get("/logging/stats/")
async def get_logging_stats():
    """Get detailed logging statistics"""
    return violation_logger.get_stats()

@app.post("/logging/cooldown/")
async def set_logging_cooldown(cooldown_seconds: int = Body(..., embed=True)):
    """Set the cooldown period between violation logs"""
    try:
        if cooldown_seconds < 1:
            raise HTTPException(status_code=400, detail="Cooldown must be at least 1 second")
        if cooldown_seconds > 300:
            raise HTTPException(status_code=400, detail="Cooldown cannot exceed 300 seconds (5 minutes)")
        
        violation_logger.set_cooldown(cooldown_seconds)
        return {
            "success": True,
            "cooldown_seconds": cooldown_seconds,
            "message": f"Cooldown set to {cooldown_seconds} seconds"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error setting cooldown: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# Face Detection Toggle Endpoints
# ============================================================================

@app.post("/face-detection/toggle/")
async def toggle_face_detection():
    """Toggle face detection on/off"""
    global face_detection_enabled
    try:
        face_detection_enabled = not face_detection_enabled
        logger.info(f"Face detection {'enabled' if face_detection_enabled else 'disabled'}")
        return {
            "success": True,
            "face_detection_enabled": face_detection_enabled,
            "message": f"Face detection {'enabled' if face_detection_enabled else 'disabled'}"
        }
    except Exception as e:
        logger.error(f"Error toggling face detection: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/face-detection/status/")
async def get_face_detection_status():
    """Get current face detection status"""
    return {
        "face_detection_enabled": face_detection_enabled
    }

@app.post("/logging/enable/")
async def enable_logging():
    """Enable violation logging"""
    violation_logger.enable_logging()
    return {
        "success": True,
        "logging_enabled": True,
        "message": "Logging enabled"
    }

@app.post("/logging/disable/")
async def disable_logging():
    """Disable violation logging"""
    violation_logger.disable_logging()
    return {
        "success": True,
        "logging_enabled": False,
        "message": "Logging disabled"
    }

@app.get("/system/status/")
async def get_system_status():
    """Get real-time system status information"""
    import psutil
    import json
    from datetime import date
    
    try:
        # Get model info
        model_info = {
            "name": detector.model_name if hasattr(detector, 'model_name') else "YOLO",
            "device": "GPU" if detector.device == "cuda" else "CPU",
            "status": "Active" if detector else "Inactive"
        }
        
        # Get logging stats
        log_stats = violation_logger.get_stats()
        
        # Count today's violations
        today_count = 0
        total_count = 0
        log_folder = violation_logger.log_folder
        
        if os.path.exists(log_folder):
            today_date = date.today().strftime("%Y%m%d")
            all_files = [f for f in os.listdir(log_folder) if f.endswith('.jpg')]
            total_count = len(all_files)
            
            for filename in all_files:
                if filename.startswith("violation_"):
                    try:
                        file_date = filename.split("_")[1]
                        if file_date == today_date:
                            today_count += 1
                    except:
                        continue
        
        # Get student database count
        student_count = 0
        students_file = "students.json"
        if os.path.exists(students_file):
            try:
                with open(students_file, 'r') as f:
                    students_db = json.load(f)
                    student_count = len(students_db)
            except:
                pass
        
        # Get system metrics
        cpu_percent = psutil.cpu_percent(interval=0.1)
        memory = psutil.virtual_memory()
        
        # Get GPU info if available
        gpu_info = None
        if detector.device == "cuda":
            try:
                import torch
                if torch.cuda.is_available():
                    gpu_info = {
                        "name": torch.cuda.get_device_name(0),
                        "memory_allocated": f"{torch.cuda.memory_allocated(0) / 1024**3:.2f} GB",
                        "memory_reserved": f"{torch.cuda.memory_reserved(0) / 1024**3:.2f} GB",
                        "memory_total": f"{torch.cuda.get_device_properties(0).total_memory / 1024**3:.2f} GB"
                    }
            except:
                pass
        
        # Webcam status
        webcam_status = "Active" if webcam_active else "Stopped"
        
        return {
            "model": model_info,
            "logging": {
                "enabled": log_stats.get("logging_enabled", False),
                "today_violations": today_count,
                "total_violations": total_count,
                "cooldown": log_stats.get("cooldown_seconds", 0)
            },
            "database": {
                "students_enrolled": student_count
            },
            "system": {
                "cpu_usage": f"{cpu_percent:.1f}%",
                "memory_usage": f"{memory.percent:.1f}%",
                "memory_available": f"{memory.available / 1024**3:.1f} GB"
            },
            "gpu": gpu_info,
            "webcam": {
                "status": webcam_status,
                "selected_index": selected_camera_index if webcam_active else None
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting system status: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# Dashboard Endpoints - Get Logs, Filter, Delete
# ============================================================================

@app.get("/dashboard/debug/logged-today/")
async def debug_logged_today():
    """Debug endpoint to see logged_today state"""
    return {
        "logged_today": violation_logger.logged_today,
        "current_date": datetime.now().strftime("%Y-%m-%d"),
        "log_folder": violation_logger.log_folder,
        "logging_enabled": violation_logger.is_logging_enabled()
    }

@app.get("/dashboard/logs/")
async def get_logs(date: Optional[str] = None, model: Optional[str] = None, page: int = 1, per_page: int = 10):
    """
    Get violation logs with pagination and optional date/model filtering.
    
    Args:
        date: Optional date filter in YYYY-MM-DD format. If None, shows today's logs.
        model: Optional model filter. If None or 'all', shows all models.
        page: Page number (1-indexed)
        per_page: Number of items per page
    
    Returns:
        {
            "logs": [...],
            "total": int,
            "page": int,
            "per_page": int,
            "total_pages": int,
            "date": str,
            "model": str
        }
    """
    try:
        # Default to today if no date specified
        if not date:
            date = datetime.now().strftime("%Y-%m-%d")
        
        # Get all log files
        log_folder = violation_logger.log_folder
        if not os.path.exists(log_folder):
            return {
                "logs": [],
                "total": 0,
                "page": page,
                "per_page": per_page,
                "total_pages": 0,
                "date": date
            }
        
        # Get all image files
        all_files = [f for f in os.listdir(log_folder) if f.endswith('.jpg')]
        logger.info(f"Dashboard: Found {len(all_files)} total image files in {log_folder}")
        
        # Filter by date
        filtered_files = []
        target_date_prefix = date.replace("-", "")  # Convert 2025-10-28 to 20251028
        logger.info(f"Dashboard: Filtering for date {date} (prefix: {target_date_prefix})")
        
        for filename in all_files:
            # Extract date from filename (format: violation_YYYYMMDD_HHMMSS_mmm.jpg)
            if filename.startswith("violation_"):
                try:
                    file_date = filename.split("_")[1]  # Get YYYYMMDD part
                    if file_date == target_date_prefix:
                        filtered_files.append(filename)
                        logger.debug(f"Dashboard: Matched file {filename}")
                except Exception as e:
                    logger.debug(f"Dashboard: Failed to parse filename {filename}: {e}")
                    continue
        
        logger.info(f"Dashboard: {len(filtered_files)} files matched date {date}")
        
        # Sort by timestamp (newest first)
        filtered_files.sort(reverse=True)
        
        # Load daily logs to get person names and violations
        daily_logs = violation_logger.logged_today
        logger.info(f"Dashboard: logged_today has {len(daily_logs)} entries")
        
        # Build metadata for all files and apply model filter BEFORE pagination
        all_logs_metadata = []
        for filename in filtered_files:
            filepath = os.path.join(log_folder, filename)
            # Normalize path for comparison (handle mixed separators)
            filepath_normalized = os.path.normpath(filepath)
            
            # Extract timestamp from filename
            try:
                parts = filename.replace(".jpg", "").split("_")
                date_part = parts[1]  # YYYYMMDD
                time_part = parts[2]  # HHMMSS
                ms_part = parts[3] if len(parts) > 3 else "000"  # mmm
                
                # Format timestamp
                timestamp_str = f"{date_part[:4]}-{date_part[4:6]}-{date_part[6:8]} {time_part[:2]}:{time_part[2:4]}:{time_part[4:6]}.{ms_part}"
                
                # Find matching file in daily logs (new structure uses filepath as key)
                person_name = "Unknown"
                violations = []
                license_plates = []
                model_name = "Unknown"
                matched = False
                
                # Try direct lookup first (new structure)
                if filepath_normalized in daily_logs:
                    log_info = daily_logs[filepath_normalized]
                    person_name = log_info.get("person", "Unknown")
                    violations = log_info.get("items", [])
                    license_plates = log_info.get("license_plates", [])
                    model_name = log_info.get("model", "Unknown")
                    matched = True
                    logger.debug(f"Dashboard: Matched {filename} to person {person_name}, model {model_name}, plates: {license_plates}")
                else:
                    # Fallback: search through all entries (legacy format compatibility)
                    for key, log_info in daily_logs.items():
                        logged_filepath = os.path.normpath(log_info.get("filepath", ""))
                        if logged_filepath == filepath_normalized:
                            person_name = log_info.get("person", key)  # key might be person name in old format
                            violations = log_info.get("items", [])
                            license_plates = log_info.get("license_plates", [])
                            model_name = log_info.get("model", "Unknown")
                            matched = True
                            logger.debug(f"Dashboard: Matched {filename} (legacy) to person {person_name}, model {model_name}")
                            break
                
                if not matched:
                    logger.warning(f"Dashboard: No match in logged_today for {filename} (filepath: {filepath_normalized})")
                    logger.info(f"Dashboard: Including {filename} anyway with basic info")
                
                # Apply model filter if specified
                if model and model.lower() != 'all' and model_name.lower() != model.lower():
                    logger.debug(f"Dashboard: Skipping {filename} - model '{model_name}' doesn't match filter '{model}'")
                    continue
                
                all_logs_metadata.append({
                    "id": filename,
                    "filename": filename,
                    "timestamp": timestamp_str,
                    "person": person_name,
                    "violations": violations if violations else ["Violation (details unavailable)"],
                    "license_plates": license_plates,
                    "model": model_name,
                    "image_url": f"/api/dashboard/image/{filename}"
                })
            except Exception as e:
                logger.error(f"Dashboard: Error processing file {filename}: {e}", exc_info=True)
                continue
        
        # Apply pagination AFTER filtering
        total = len(all_logs_metadata)
        total_pages = (total + per_page - 1) // per_page if total > 0 else 0
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        logs = all_logs_metadata[start_idx:end_idx]
        
        logger.info(f"Dashboard: Returning {len(logs)} of {total} total logs for date {date}, model filter: {model or 'all'}")
        
        return {
            "logs": logs,
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": total_pages,
            "date": date,
            "model": model or "all"
        }
        
    except Exception as e:
        logger.error(f"Error getting logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard/image/{filename}")
async def get_log_image(filename: str):
    """
    Serve a specific log image file.
    
    Args:
        filename: Name of the image file
    
    Returns:
        FileResponse with the image
    """
    try:
        filepath = os.path.join(violation_logger.log_folder, filename)
        
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="Image not found")
        
        # Security check - ensure filename doesn't contain path traversal
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        from fastapi.responses import FileResponse
        return FileResponse(filepath, media_type="image/jpeg")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving image {filename}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/dashboard/log/{filename}")
async def delete_log(filename: str):
    """
    Delete a specific log entry and its image.
    
    Args:
        filename: Name of the log file to delete
    
    Returns:
        Success message
    """
    try:
        filepath = os.path.join(violation_logger.log_folder, filename)
        
        # Security check
        if ".." in filename or "/" in filename or "\\" in filename:
            raise HTTPException(status_code=400, detail="Invalid filename")
        
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="Log not found")
        
        # Delete the image file
        os.remove(filepath)
        
        # Remove from daily logs if present
        daily_logs = violation_logger.logged_today.copy()
        for person, log_info in list(daily_logs.items()):
            if log_info.get("filepath") == filepath:
                del violation_logger.logged_today[person]
                violation_logger._save_daily_logs()
                break
        
        logger.info(f"Deleted log: {filename}")
        return {
            "success": True,
            "message": f"Log {filename} deleted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting log {filename}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/dashboard/logs/clear/{date}")
async def clear_logs_by_date(date: str):
    """
    Clear all logs for a specific date.
    
    Args:
        date: Date in YYYY-MM-DD format
    
    Returns:
        Count of deleted logs
    """
    try:
        log_folder = violation_logger.log_folder
        if not os.path.exists(log_folder):
            return {
                "success": True,
                "deleted_count": 0,
                "message": "No logs found"
            }
        
        # Get all image files for this date
        target_date_prefix = date.replace("-", "")  # Convert 2025-10-28 to 20251028
        deleted_count = 0
        
        for filename in os.listdir(log_folder):
            if filename.startswith("violation_") and filename.endswith('.jpg'):
                try:
                    file_date = filename.split("_")[1]  # Get YYYYMMDD part
                    if file_date == target_date_prefix:
                        filepath = os.path.join(log_folder, filename)
                        os.remove(filepath)
                        deleted_count += 1
                        
                        # Remove from daily logs
                        for person, log_info in list(violation_logger.logged_today.items()):
                            if log_info.get("filepath") == filepath:
                                del violation_logger.logged_today[person]
                except Exception as e:
                    logger.error(f"Error deleting file {filename}: {e}")
                    continue
        
        # Save updated daily logs
        if deleted_count > 0:
            violation_logger._save_daily_logs()
        
        logger.info(f"Cleared {deleted_count} logs for date {date}")
        return {
            "success": True,
            "deleted_count": deleted_count,
            "message": f"Deleted {deleted_count} logs for {date}"
        }
        
    except Exception as e:
        logger.error(f"Error clearing logs for date {date}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard/dates/")
async def get_available_dates():
    """
    Get list of dates that have logs available.
    
    Returns:
        List of dates in YYYY-MM-DD format
    """
    try:
        log_folder = violation_logger.log_folder
        if not os.path.exists(log_folder):
            return {"dates": []}
        
        # Get all unique dates from log files
        dates_set = set()
        
        for filename in os.listdir(log_folder):
            if filename.startswith("violation_") and filename.endswith('.jpg'):
                try:
                    file_date = filename.split("_")[1]  # Get YYYYMMDD part
                    # Convert to YYYY-MM-DD format
                    formatted_date = f"{file_date[:4]}-{file_date[4:6]}-{file_date[6:8]}"
                    dates_set.add(formatted_date)
                except:
                    continue
        
        # Sort dates (newest first)
        dates = sorted(list(dates_set), reverse=True)
        
        return {"dates": dates}
        
    except Exception as e:
        logger.error(f"Error getting available dates: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard/models/")
async def get_available_models():
    """
    Get list of all available detection models in the system.
    
    Returns:
        List of model names available for filtering
    """
    try:
        # Get all available models from the model discovery system
        models_set = set()
        
        if detector and hasattr(detector, 'model_discovery'):
            available_models = detector.model_discovery.get_all_models()
            for model_name in available_models.keys():
                models_set.add(model_name)
        
        # Sort models alphabetically
        models = sorted(list(models_set))
        
        logger.info(f"Available models for filtering: {models}")
        
        return {"models": models}
        
    except Exception as e:
        logger.error(f"Error getting available models: {e}")
        raise HTTPException(status_code=500, detail=str(e))

def _generate_excel_report(date: str, model: Optional[str] = None, save_to_file: bool = False):
    """
    Internal function to generate Excel report
    
    Args:
        date: Date in YYYY-MM-DD format
        model: Optional model name to filter by
        save_to_file: If True, saves to file and returns path. If False, returns BytesIO buffer
        
    Returns:
        Tuple of (workbook_buffer_or_path, violation_count)
    """
    import json
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    
    # Load student database
    students_file = "students.json"
    students_db = {}
    if os.path.exists(students_file):
        try:
            with open(students_file, 'r') as f:
                students_db = json.load(f)
        except Exception as e:
            logger.error(f"Error loading students.json: {e}")
    
    # Get logs for this date
    log_folder = violation_logger.log_folder
    if not os.path.exists(log_folder):
        raise HTTPException(status_code=404, detail="No logs found")
    
    # Filter files by date
    target_date_prefix = date.replace("-", "")
    violations = []
    
    for filename in os.listdir(log_folder):
        if filename.startswith("violation_") and filename.endswith('.jpg'):
            try:
                file_date = filename.split("_")[1]
                if file_date == target_date_prefix:
                    # Extract timestamp
                    time_part = filename.split("_")[2]
                    timestamp = f"{file_date[:4]}-{file_date[4:6]}-{file_date[6:8]} {time_part[:2]}:{time_part[2:4]}:{time_part[4:6]}"
                    
                    # Find person and violations from daily logs
                    filepath = os.path.join(log_folder, filename)
                    filepath_normalized = os.path.normpath(filepath)
                    person_name = "Unknown"
                    items = []
                    license_plates = []
                    model_name = "Unknown"
                    
                    # logged_today uses filepath as key in new format
                    if filepath_normalized in violation_logger.logged_today:
                        log_info = violation_logger.logged_today[filepath_normalized]
                        person_name = log_info.get("person", "Unknown")
                        items = log_info.get("items", [])
                        license_plates = log_info.get("license_plates", [])
                        model_name = log_info.get("model", "Unknown")
                    else:
                        # Fallback: search through all entries for matching filepath
                        for key, log_info in violation_logger.logged_today.items():
                            if not isinstance(log_info, dict):
                                continue
                            logged_filepath = os.path.normpath(log_info.get("filepath", key))
                            if logged_filepath == filepath_normalized:
                                person_name = log_info.get("person", "Unknown")
                                items = log_info.get("items", [])
                                license_plates = log_info.get("license_plates", [])
                                model_name = log_info.get("model", "Unknown")
                                break
                    
                    # Apply model filter if specified
                    if model and model.lower() != 'all' and model_name.lower() != model.lower():
                        continue
                    
                    # Get student details
                    student_info = students_db.get(person_name, {})
                    
                    violations.append({
                        'person': person_name,
                        'full_name': student_info.get('full_name', person_name),
                        'usn': student_info.get('usn', 'N/A'),
                        'department': student_info.get('department', 'N/A'),
                        'branch': student_info.get('branch', 'N/A'),
                        'email': student_info.get('email', 'N/A'),
                        'timestamp': timestamp,
                        'violations': ', '.join(items) if items else 'N/A',
                        'license_plates': ', '.join(license_plates) if license_plates else 'N/A',
                        'model': model_name,
                        'image': filename
                    })
            except Exception as e:
                logger.error(f"Error processing file {filename}: {e}", exc_info=True)
                continue
    
    model_filter_msg = f" for model '{model}'" if model and model.lower() != 'all' else ""
    if not violations:
        raise HTTPException(status_code=404, detail=f"No violations found for {date}{model_filter_msg}")
    
    # Sort by timestamp
    violations.sort(key=lambda x: x['timestamp'])
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Non-Compliance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    model_suffix = f" - {model}" if model and model.lower() != 'all' else " - All Models"
    title_cell.value = f"DressGuard Non-Compliance Report - {date}{model_suffix}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Add headers
    headers = ['#', 'Full Name', 'USN', 'Department', 'Branch', 'Email', 'Timestamp', 'Violations', 'License Plates', 'Model', 'Image']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    ws.row_dimensions[2].height = 25
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 35  # Wider for images
    
    # Add data rows with embedded images
    for idx, violation in enumerate(violations, 1):
        row_num = idx + 2
        row_data = [
            idx,
            violation['full_name'],
            violation['usn'],
            violation['department'],
            violation['branch'],
            violation['email'],
            violation['timestamp'],
            violation['violations'],
            violation['license_plates'],
            violation['model']
        ]
        
        # Add text data
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = border
        
        # Add embedded image in column K (11)
        try:
            image_path = os.path.join(log_folder, violation['image'])
            if os.path.exists(image_path):
                # Open and resize image for WhatsApp compatibility with good quality
                pil_img = PILImage.open(image_path)
                orig_width, orig_height = pil_img.size
                
                # Balanced dimensions - clear but not too large
                target_width = 150
                aspect_ratio = orig_height / orig_width
                target_height = int(target_width * aspect_ratio)
                
                # Limit height to max 115 pixels
                max_height = 115
                if target_height > max_height:
                    target_height = max_height
                    target_width = int(max_height / aspect_ratio)
                
                # Resize image with high-quality downsampling
                resized_img = pil_img.resize((target_width, target_height), PILImage.Resampling.LANCZOS)
                
                # Convert to RGB if needed (removes alpha channel)
                if resized_img.mode in ('RGBA', 'LA', 'P'):
                    rgb_img = PILImage.new('RGB', resized_img.size, (255, 255, 255))
                    if resized_img.mode == 'P':
                        resized_img = resized_img.convert('RGBA')
                    rgb_img.paste(resized_img, mask=resized_img.split()[-1] if resized_img.mode == 'RGBA' else None)
                    resized_img = rgb_img
                
                # Save to temporary buffer with JPEG compression - higher quality
                temp_buffer = BytesIO()
                resized_img.save(temp_buffer, format='JPEG', quality=85, optimize=True)
                temp_buffer.seek(0)
                
                # Create Excel image from buffer
                img = XLImage(temp_buffer)
                img.width = target_width
                img.height = target_height
                
                # Position image in cell K (column 11)
                img.anchor = f'K{row_num}'
                ws.add_image(img)
                
                # Set row height to accommodate image (convert pixels to points: pixels * 0.75 + padding)
                ws.row_dimensions[row_num].height = target_height * 0.75 + 5
            else:
                # If image doesn't exist, just show filename
                cell = ws.cell(row=row_num, column=11)
                cell.value = violation['image']
                cell.alignment = cell_alignment
                cell.border = border
                ws.row_dimensions[row_num].height = 20
        except Exception as e:
            logger.error(f"Error embedding image {violation['image']}: {e}", exc_info=True)
            # Fallback to filename if image embedding fails
            cell = ws.cell(row=row_num, column=11)
            cell.value = violation['image']
            cell.alignment = cell_alignment
            cell.border = border
            ws.row_dimensions[row_num].height = 20
    
    # Add summary at bottom
    summary_row = len(violations) + 4
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = f"Total Violations: {len(violations)}"
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Save to file or return buffer
    if save_to_file:
        report_filename = f"DressGuard_Report_{date.replace('-', '')}.xlsx"
        report_path = os.path.join("non_compliance_logs", report_filename)
        os.makedirs("non_compliance_logs", exist_ok=True)
        wb.save(report_path)
        return report_path, len(violations)
    else:
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer, len(violations)

@app.get("/dashboard/report/{date}")
async def generate_report(date: str, model: Optional[str] = None):
    """
    Generate Excel report for a specific date and optionally filtered by model.
    
    Args:
        date: Date in YYYY-MM-DD format
        model: Optional model name to filter by (e.g., 'Vehicle Helmet', 'Student Uniform')
    
    Returns:
        Excel file download with violation details
    """
    try:
        from fastapi.responses import StreamingResponse
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="openpyxl not installed. Run: pip install openpyxl"
            )
        
        # Use shared report generation function
        excel_buffer, violation_count = _generate_excel_report(date, model, save_to_file=False)
        
        # Generate filename
        safe_date = date.replace("-", "")
        filename = f"DressGuard_Report_{safe_date}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating report for {date}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/dashboard/whatsapp/status/")
async def get_whatsapp_status():
    """Check if WhatsApp Web integration is enabled and configured"""
    try:
        import config
        is_enabled = whatsapp_sender.is_enabled()
        # WhatsApp Web doesn't need Twilio - just check if service is running
        return {
            "enabled": is_enabled,
            "configured": is_enabled,  # If enabled, it's configured (WhatsApp Web connected)
            "recipients": getattr(config, 'WHATSAPP_RECIPIENTS', [])
        }
    except Exception as e:
        logger.error(f"Error checking WhatsApp status: {e}")
        return {
            "enabled": False,
            "configured": False,
            "error": str(e)
        }

@app.post("/dashboard/send-whatsapp/")
async def send_report_to_whatsapp(date: str = Body(...), 
                                   model: Optional[str] = Body(None),
                                   recipients: Optional[List[str]] = Body(None)):
    """
    Generate Excel report and send via WhatsApp with file attachment
    
    Args:
        date: Report date in YYYY-MM-DD format
        model: Optional model name to filter by
        recipients: Optional list of recipient phone numbers (uses config default if not provided)
    
    Returns:
        Status of WhatsApp message sending
    """
    try:
        import config
        
        # Check if WhatsApp is enabled
        if not whatsapp_sender.is_enabled():
            raise HTTPException(
                status_code=503, 
                detail="WhatsApp Web service not ready. Please start: cd whatsapp-service && node server.js"
            )
        
        # Get recipients from config if not provided
        if not recipients:
            recipients = getattr(config, 'WHATSAPP_RECIPIENTS', [])
        
        if not recipients:
            raise HTTPException(
                status_code=400,
                detail="No recipients specified. Add phone numbers to WHATSAPP_RECIPIENTS in config.py"
            )
        
        logger.info(f"Generating report for {date}...")
        
        # Use shared report generation function (save to file for WhatsApp)
        report_path, violation_count = _generate_excel_report(date, model, save_to_file=True)
        logger.info(f"Report generated: {report_path}")
        
        # Send via WhatsApp with attachment
        results = []
        for recipient in recipients:
            result = whatsapp_sender.send_report_with_file(
                to_number=recipient,
                date=date,
                total_violations=violation_count,
                report_path=report_path
            )
            results.append({
                "recipient": recipient,
                "success": result.get('success', False),
                "message_sid": result.get('message_sid'),
                "error": result.get('error')
            })
        
        success_count = sum(1 for r in results if r['success'])
        
        return {
            "success": success_count > 0,
            "date": date,
            "violation_count": violation_count,
            "total_recipients": len(recipients),
            "successful_sends": success_count,
            "report_generated": True,
            "report_path": report_path,
            "results": results
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error sending WhatsApp report for {date}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# Webcam Stream Endpoints
# ============================================================================

# Global variable to control webcam stream
webcam_active = False
webcam_cap = None
selected_camera_index = 0  # Default to camera 0
multiple_people_warning_active = False  # Global flag for notification
multiple_people_warning_timestamp = 0  # When the warning was triggered

@app.get("/webcam/warning/")
async def get_webcam_warning():
    """Check if there's an active multiple people warning"""
    global multiple_people_warning_active, multiple_people_warning_timestamp
    
    current_time = time.time()
    
    # Check if warning has expired (3 seconds)
    if multiple_people_warning_active and (current_time - multiple_people_warning_timestamp) > 3.0:
        multiple_people_warning_active = False
    
    return {
        "warning_active": multiple_people_warning_active,
        "message": "Multiple people detected - Only one person should be in frame" if multiple_people_warning_active else None
    }

@app.get("/cameras/list/")
async def list_cameras():
    """List available cameras on the system"""
    available_cameras = []
    
    # Try to detect up to 10 cameras
    for i in range(10):
        cap = cv2.VideoCapture(i)
        if cap.isOpened():
            # Get camera name/info if available
            ret, frame = cap.read()
            if ret:
                available_cameras.append({
                    "index": i,
                    "name": f"Camera {i}",
                    "resolution": f"{int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))}x{int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))}"
                })
            cap.release()
    
    logger.info(f"Found {len(available_cameras)} cameras")
    return {"cameras": available_cameras}

@app.post("/cameras/select/")
async def select_camera(camera_index: int = Body(..., embed=True)):
    """Select which camera to use for streaming"""
    global selected_camera_index, webcam_active, webcam_cap
    
    # Validate camera index
    test_cap = cv2.VideoCapture(camera_index)
    if not test_cap.isOpened():
        test_cap.release()
        raise HTTPException(status_code=400, detail=f"Camera {camera_index} not available")
    test_cap.release()
    
    # Stop current stream if active
    if webcam_active:
        webcam_active = False
        if webcam_cap is not None:
            webcam_cap.release()
            webcam_cap = None
        await asyncio.sleep(0.5)  # Give time for stream to stop
    
    # Update selected camera
    selected_camera_index = camera_index
    logger.info(f"Selected camera index: {camera_index}")
    
    return {
        "success": True, 
        "message": f"Camera {camera_index} selected",
        "camera_index": camera_index
    }

def generate_webcam_frames():
    """Generate MJPEG frames from webcam with YOLO detection and distance checking"""
    global webcam_cap, webcam_active, selected_camera_index
    
    # Initialize webcam with selected camera index
    webcam_cap = cv2.VideoCapture(selected_camera_index)
    
    if not webcam_cap.isOpened():
        logger.error("Could not open webcam")
        yield b''
        return
    
    # Set webcam properties for better performance
    webcam_cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    webcam_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    webcam_cap.set(cv2.CAP_PROP_FPS, 30)
    webcam_cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # Reduce buffer to minimize lag
    
    logger.info("Webcam stream started")
    webcam_active = True
    
    # Frame processing control
    frame_count = 0
    logged_persons = set()  # Track logged persons in current session
    last_face_detection_time = 0
    face_detection_interval = 7.0  # Run face detection every 5 seconds
    current_status = "LOADING"  # Track current operation status
    last_logged_time = {}  # Track when each person was last logged {name: timestamp}
    session_reset_timeout = 15.0  # Reset session tracking after 15 seconds of inactivity
    multiple_people_warning_time = 0  # Track when multiple people warning was shown
    multiple_people_warning_duration = 3.0  # Show warning for 3 seconds
    
    # JPEG encoding parameters for better performance
    encode_params = [
        cv2.IMWRITE_JPEG_QUALITY, WEBCAM_JPEG_QUALITY,
        cv2.IMWRITE_JPEG_OPTIMIZE, 1    # Optimize encoding
    ]
    
    # FPS control
    frame_delay = 1.0 / WEBCAM_FPS_LIMIT if WEBCAM_FPS_LIMIT > 0 else 0.01
    
    try:
        while webcam_active:
            ret, frame = webcam_cap.read()
            
            if not ret:
                logger.warning("Failed to grab webcam frame")
                break
            
            # Frame skipping for performance (optional)
            if WEBCAM_SKIP_FRAMES > 0 and frame_count % (WEBCAM_SKIP_FRAMES + 1) != 0:
                frame_count += 1
                continue
            
            try:
                frame_count += 1
                current_time = time.time()
                annotated_frame = frame.copy()
                
                # Reset session tracking if person hasn't been seen for a while
                expired_persons = []
                for person_name, last_time in last_logged_time.items():
                    if current_time - last_time > session_reset_timeout:
                        expired_persons.append(person_name)
                
                for person_name in expired_persons:
                    if person_name in logged_persons:
                        logged_persons.remove(person_name)
                        del last_logged_time[person_name]
                        logger.info(f"Session reset for {person_name} - can be logged again if they return")
                
                # Set initial status
                current_status = "DETECTING"
                
                # Always run YOLO detection
                results = detector.detect(frame, confidence_threshold=0.6)
                is_compliant, non_compliant_items, compliance_details = compliance_manager.check_compliance(results)
                
                # If non-compliant and logging is enabled, process violation logging
                if not is_compliant and violation_logger.is_logging_enabled():
                    # Run face detection/logging every 5 seconds
                    if current_time - last_face_detection_time >= face_detection_interval:
                        last_face_detection_time = current_time
                        face_results = []
                        
                        # If face detection is enabled, perform face recognition
                        if face_detection_enabled:
                            current_status = "SCANNING FACE..."
                            # Convert frame to RGB for face detection
                            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                            face_results = detect_and_identify_faces(frame_rgb)
                            
                            # If no faces detected, mark as Unknown
                            if not face_results:
                                logger.info("No faces detected in video frame - marking as Unknown")
                                face_results = [{'name': 'Unknown', 'confidence': 0, 'user_id': None}]
                        else:
                            # Face detection disabled - mark as Unknown for logging
                            current_status = "LOGGING VIOLATION..."
                            logger.info("Face detection disabled - marking as Unknown for logging")
                            face_results = [{'name': 'Unknown', 'confidence': 0, 'user_id': None}]
                        
                        # Handle logging based on face detection results
                        # Debug: Log what faces were detected
                        detected_names = [f"{face.get('name')} ({face.get('confidence', 0):.1f}%)" for face in face_results]
                        logger.info(f"Face detection results: {detected_names}")
                        
                        # Check if multiple people detected in frame (only when face detection enabled)
                        if face_detection_enabled and len(face_results) > 1:
                            global multiple_people_warning_active, multiple_people_warning_timestamp
                            logger.warning(f"Multiple people detected in frame: {detected_names}")
                            current_status = "⚠️ MULTIPLE PEOPLE DETECTED - Only one person should be in frame"
                            multiple_people_warning_time = current_time
                            # Set global warning flag for frontend notification
                            multiple_people_warning_active = True
                            multiple_people_warning_timestamp = current_time
                            # Don't process or log anything when multiple people detected
                        else:
                            # Single person or face detection disabled - proceed with logging
                            known_faces = [face for face in face_results if face.get('name') != 'Unknown']
                            
                            # Process first known face only (one person at a time)
                            if known_faces:
                                    face_info = known_faces[0]
                                    name = face_info.get('name')
                                    confidence = face_info.get('confidence', 0)
                                    
                                    logger.info(f"Processing known face: {name} with confidence {confidence:.1f}%")
                                    
                                    # If Unknown was logged but now we identified a known person, delete Unknown immediately
                                    # This happens regardless of whether we log the known person or not
                                    if 'Unknown' in logged_persons or violation_logger._is_person_logged_today('Unknown'):
                                        logger.info(f"Known person {name} detected - deleting any Unknown logs (session cleanup)")
                                        violation_logger._delete_previous_log('Unknown')
                                        if 'Unknown' in logged_persons:
                                            logged_persons.remove('Unknown')
                                        if 'Unknown' in last_logged_time:
                                            del last_logged_time['Unknown']
                                    
                                    # Check session tracking - only skip if logged very recently (within session timeout)
                                    should_skip = False
                                    if name in logged_persons:
                                        time_since_logged = current_time - last_logged_time.get(name, 0)
                                        if time_since_logged < session_reset_timeout:
                                            # Still within session timeout - skip logging
                                            should_skip = True
                                            time_until_reset = session_reset_timeout - time_since_logged
                                            current_status = f"{name} - Logged (resets in {int(time_until_reset)}s)"
                                            logger.debug(f"{name} already logged this session, {int(time_until_reset)}s until reset")
                                    
                                    if not should_skip:
                                        # Either first time or session timeout expired - can log/update
                                        
                                        # Check if person was logged earlier today (from database)
                                        if violation_logger._is_person_logged_today(name):
                                            current_status = f"UPDATING: {name}..."
                                            logger.info(f"Person {name} already logged today - deleting old log and replacing with new one")
                                            # Delete previous log before saving new one
                                            violation_logger._delete_previous_log(name)
                                        else:
                                            current_status = f"LOGGING: {name}..."
                                        
                                        logger.info(f"About to save violation for {name}")
                                        logger.info(f"Face results being passed to save_violation: {face_results}")
                                        logger.info(f"Non-compliant items: {non_compliant_items}")
                                        
                                        # IMPORTANT: Only pass the known face we're logging, not all face_results
                                        # This prevents Unknown faces from being saved when we detect a known person
                                        known_face_only = [face_info]  # Only the first known face
                                        logger.info(f"Filtered to known face only: {known_face_only}")
                                        
                                        # Log the violation (will replace if person was logged before)
                                        save_result = violation_logger.save_violation(
                                            frame.copy(),
                                            results,
                                            known_face_only,  # Pass only the known face, not all faces
                                            {
                                                'is_compliant': is_compliant,
                                                'non_compliant_items': non_compliant_items
                                            },
                                            current_model=detector.current_model
                                        )
                                        
                                        if save_result:
                                            logger.info(f"✓ Violation logged successfully: {name}, items: {non_compliant_items}")
                                            logged_persons.add(name)
                                            last_logged_time[name] = current_time  # Track when this person was logged
                                            current_status = f"✓ LOGGED: {name}"
                                        else:
                                            logger.warning(f"✗ Violation NOT logged for {name} - save_violation returned False")
                                            current_status = f"Failed to log: {name}"
                            else:
                                # Only unknown faces or face detection disabled - log as Unknown
                                if 'Unknown' not in logged_persons:
                                    current_status = "LOGGING: Unknown..."
                                    logger.info("Unknown person detected - logging violation")
                                    violation_logged = violation_logger.save_violation(
                                        frame.copy(),
                                        results,
                                        face_results,  # Contains Unknown face entry
                                        {
                                            'is_compliant': is_compliant,
                                            'non_compliant_items': non_compliant_items
                                        },
                                        current_model=detector.current_model
                                    )
                                    if violation_logged:
                                        logged_persons.add('Unknown')
                                        last_logged_time['Unknown'] = current_time
                                        current_status = "✓ LOGGED: Unknown"
                                        logger.info("✓ Violation logged for Unknown")
                                    else:
                                        current_status = "Failed to log Unknown"
                                        logger.warning("Failed to log violation for Unknown")
                                else:
                                    # Already logged Unknown in this session
                                    time_since_logged = current_time - last_logged_time.get('Unknown', current_time)
                                    time_until_reset = session_reset_timeout - time_since_logged
                                    if time_until_reset > 0:
                                        current_status = f"Unknown - Logged (resets in {int(time_until_reset)}s)"
                                    else:
                                        current_status = "Unknown - Already logged"
                    else:
                        # Waiting for next scan
                        time_until_next = face_detection_interval - (current_time - last_face_detection_time)
                        if len(logged_persons) > 0:
                            tracked_names = ", ".join(logged_persons)
                            current_status = f"LOGGED: {tracked_names}"
                        else:
                            current_status = f"Next scan in {int(time_until_next)}s"
                
                # Draw bounding boxes
                annotated_frame = draw_detections_on_frame(annotated_frame, results)
                
                # Dynamic status indicator with color coding - TOP LEFT with background
                status_text = current_status
                
                # Color coding based on status
                if "LOADING" in status_text:
                    status_color = (255, 165, 0)  # Orange - loading
                elif "SCANNING" in status_text:
                    status_color = (0, 255, 255)  # Cyan - scanning face
                elif "LOGGING" in status_text:
                    status_color = (255, 255, 0)  # Yellow - actively logging
                elif "✓ LOGGED" in status_text or "LOGGED:" in status_text:
                    status_color = (0, 255, 0)  # Green - successfully logged
                elif "Already logged" in status_text or "Logged today" in status_text:
                    status_color = (255, 165, 0)  # Orange - already logged
                elif "Next scan" in status_text:
                    status_color = (200, 200, 200)  # Light gray - waiting
                elif "No face" in status_text:
                    status_color = (0, 165, 255)  # Light blue - no face found
                else:
                    status_color = (0, 255, 0)  # Green - default detecting
                
                # Draw status with background for better visibility
                h, w = annotated_frame.shape[:2]
                font = cv2.FONT_HERSHEY_SIMPLEX
                font_scale = 0.6
                thickness = 2
                (text_width, text_height), baseline = cv2.getTextSize(status_text, font, font_scale, thickness)
                
                # Background rectangle - top left
                padding = 8
                bg_x1, bg_y1 = 5, 5
                bg_x2, bg_y2 = text_width + padding * 2, text_height + padding * 2
                
                overlay = annotated_frame.copy()
                cv2.rectangle(overlay, (bg_x1, bg_y1), (bg_x2, bg_y2), (0, 0, 0), -1)
                cv2.addWeighted(overlay, 0.7, annotated_frame, 0.3, 0, annotated_frame)
                
                # Status text
                cv2.putText(annotated_frame, status_text, (padding, text_height + padding),
                           font, font_scale, status_color, thickness)
                
                # Encode frame as JPEG with optimized parameters
                ret, buffer = cv2.imencode('.jpg', annotated_frame, encode_params)
                
                if not ret:
                    continue
                
                frame_bytes = buffer.tobytes()
                
                # Yield frame in MJPEG format
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
                
                # Dynamic frame rate control
                time.sleep(frame_delay)

                
            except Exception as e:
                logger.error(f"Error processing webcam frame: {e}")
                continue
                
    finally:
        if webcam_cap is not None:
            webcam_cap.release()
            webcam_cap = None
        webcam_active = False
        logger.info("Webcam stream stopped")

def draw_detections_on_frame(frame, detections, compliance_info=None):
    """Draw bounding boxes and labels on frame with compliance color coding
    
    Note: compliance_info parameter is kept for compatibility but not used for banner.
    Color coding is determined by checking compliance_manager directly.
    """
    annotated_frame = frame.copy()
    
    # Determine which classes are compliant/non-compliant for color coding
    compliant_classes = compliance_manager.compliant_classes
    non_compliant_classes = compliance_manager.non_compliant_classes
    
    for det in detections:
        class_name = det['class']
        class_name_lower = class_name.lower().strip()
        confidence = det['confidence']
        bbox = det['bbox']
        
        x1, y1, x2, y2 = map(int, bbox)
        
        # Determine color based on compliance status
        if class_name_lower in non_compliant_classes:
            color = (0, 0, 255)  # Red for non-compliant (BGR format)
            status = "NON-COMPLIANT"
        elif class_name_lower in compliant_classes:
            color = (0, 255, 0)  # Green for compliant
            status = "COMPLIANT"
        else:
            color = (255, 165, 0)  # Orange for neutral/unknown
            status = "NEUTRAL"
        
        # Draw rectangle
        cv2.rectangle(annotated_frame, (x1, y1), (x2, y2), color, 2)
        
        # Draw label with background
        label = f"{class_name}: {confidence:.2f}"
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.6
        thickness = 2
        
        # Get text size for background
        (text_width, text_height), baseline = cv2.getTextSize(label, font, font_scale, thickness)
        
        # Draw background rectangle
        cv2.rectangle(annotated_frame, 
                     (x1, y1 - text_height - 10), 
                     (x1 + text_width, y1), 
                     color, -1)
        
        # Draw text
        cv2.putText(annotated_frame, label, (x1, y1 - 5), 
                   font, font_scale, (255, 255, 255), thickness)
    
    return annotated_frame

@app.get("/webcam/stream/")
async def webcam_stream():
    """Stream webcam feed with real-time YOLO detection (MJPEG format)"""
    if detector is None:
        raise HTTPException(status_code=503, detail="Detector not initialized")
    
    return StreamingResponse(
        generate_webcam_frames(),
        media_type="multipart/x-mixed-replace; boundary=frame"
    )

@app.post("/webcam/stop/")
async def stop_webcam():
    """Stop the webcam stream"""
    global webcam_active, webcam_cap
    
    logger.info("Stop webcam request received")
    webcam_active = False
    
    # Give time for the stream to stop
    await asyncio.sleep(0.5)
    
    # Release the camera if it's still open
    if webcam_cap is not None:
        try:
            webcam_cap.release()
            logger.info("Webcam released successfully")
        except Exception as e:
            logger.error(f"Error releasing webcam: {e}")
        finally:
            webcam_cap = None
    
    return {"success": True, "message": "Webcam stream stopped"}